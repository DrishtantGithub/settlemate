# streamlit_app.py
"""
SettleMate - Streamlit app: convert .xls -> .xlsx (in-memory) then sum Fee and Transaction Amount entries.
Requirements (add to requirements.txt):
    streamlit
    pandas
    openpyxl>=3.1.0
    pyexcel
    pyexcel-xls
    pyexcel-xlsx
    numpy
"""
import streamlit as st
import pandas as pd
import io, zipfile, re
from datetime import datetime

import streamlit as st

# --- LOGIN SETUP ---
st.set_page_config(page_title="SettleMate", layout="wide")

# Simple password protection using Streamlit Secrets
def check_password():
    """Returns `True` if the user had a correct password."""
    def password_entered():
        """Check if entered password is correct."""
        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # remove password from memory
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input for password.
        st.text_input("Enter password:", type="password", on_change=password_entered, key="password")
        st.stop()
    elif not st.session_state["password_correct"]:
        # Password incorrect, show input + error.
        st.text_input("Enter password:", type="password", on_change=password_entered, key="password")
        st.error("😕 Password incorrect")
        st.stop()

check_password()


st.set_page_config(page_title="SettleMate - UPI settlement summarizer", layout="wide")
st.title("SettleMate — UPI Settlement Summarizer")
st.caption("Upload a ZIP of Excel files (or multiple .xlsx/.xls files). .xls files will be converted to .xlsx automatically before processing.")

# ----------------- Utilities & conversion -----------------
date_regex = re.compile(r'(\d{4}-\d{2}-\d{2})|(\d{2}-\d{2}-\d{4})|([A-Za-z]{3,9}\s+\d{1,2},\s*\d{4})')

def convert_xls_to_xlsx_bytes(xls_bytes):
    """
    Convert .xls bytes to .xlsx bytes using pyexcel -> pandas -> openpyxl writer.
    Returns bytes of an .xlsx file. Raises RuntimeError on failure.
    """
    try:
        # Lazy import (may not be installed in some environments)
        import pyexcel
        from io import BytesIO
        import pandas as pd
    except Exception as e:
        raise RuntimeError("Conversion requires pyexcel & pandas installed. Install 'pyexcel pyexcel-xls pyexcel-xlsx'. Err: " + str(e))

    try:
        book = pyexcel.get_book(file_content=xls_bytes)
        output = BytesIO()
        # Write each sheet as-is to an xlsx bytes buffer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet in book:
                # sheet.to_array() gives a 2D list
                arr = sheet.to_array()
                df = pd.DataFrame(arr)
                # write without header (keeps original layout)
                # use sheet.name if available else generate one
                sheet_name = sheet.name if getattr(sheet, 'name', None) else 'Sheet1'
                # Excel sheet names max length 31; enforce a safe length
                sheet_name = str(sheet_name)[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        return output.getvalue()
    except Exception as e:
        raise RuntimeError("Failed converting .xls to .xlsx: " + str(e))

def read_xlsx_via_openpyxl(bts):
    """Fallback loader using openpyxl directly if pandas engine fails."""
    try:
        from openpyxl import load_workbook
        from io import BytesIO
    except Exception as e:
        raise RuntimeError("openpyxl import failed; install/upgrade openpyxl. Err: " + str(e))
    wb = load_workbook(filename=BytesIO(bts), data_only=True, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        df = pd.DataFrame(rows)
        sheets[name] = df
    return sheets

def read_excel_bytes_safe(bts, fname):
    """
    Read bytes into dict of DataFrames. For .xlsx use pandas+openpyxl (preferred),
    fallback to openpyxl loader. .xls bytes should have been converted to xlsx earlier.
    """
    fname_low = fname.lower()
    ext = fname_low.split('.')[-1] if '.' in fname_low else ''
    try:
        # for .xls we expect conversion upstream; if ext == 'xls' try pandas xlrd if available
        if ext == 'xls':
            # try pandas (if xlrd installed)
            return pd.read_excel(io.BytesIO(bts), sheet_name=None, engine='xlrd', header=None)
        elif ext in ('xlsx','xlsm','xltx','xltm'):
            return pd.read_excel(io.BytesIO(bts), sheet_name=None, engine='openpyxl', header=None)
        else:
            return pd.read_excel(io.BytesIO(bts), sheet_name=None, header=None)
    except Exception as e:
        msg = str(e).lower()
        # if pandas complains about openpyxl version etc -> try fallback openpyxl
        if 'openpyxl' in msg or ('requires' in msg and 'openpyxl' in msg) or 'failed to import' in msg:
            try:
                return read_xlsx_via_openpyxl(bts)
            except Exception as fallback_e:
                raise RuntimeError(f"Pandas openpyxl error and fallback failed: {fallback_e}")
        # if xlrd missing for .xls give clear instruction
        if 'xlrd' in msg and ext == 'xls':
            raise RuntimeError("Reading .xls requires xlrd (or convert .xls to .xlsx). Install xlrd==1.2.0 or rely on automatic conversion.")
        raise

# ----------------- ZIP extraction (with conversion) -----------------
def extract_excel_files_from_zip(zip_bytes):
    """
    Return list of (name, bytes) for legit Excel files inside zip.
    Converts .xls files to .xlsx bytes before returning.
    """
    excel_files = []
    try:
        z = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except Exception as e:
        st.error(f"Uploaded zip is invalid: {e}")
        return excel_files

    for name in z.namelist():
        base = name.split('/')[-1]
        if not base:
            continue
        # skip macOS metadata / temp / lock files
        if name.startswith('__MACOSX/') or base.startswith('._') or base.startswith('~$'):
            continue
        lower = base.lower()
        if lower.endswith(('.xlsx', '.xls', '.xlsm', '.xltx', '.xltm')):
            try:
                bts = z.read(name)
                # If .xls convert to xlsx bytes
                if lower.endswith('.xls'):
                    try:
                        converted = convert_xls_to_xlsx_bytes(bts)
                        # Give converted file a new name with .xlsx extension
                        new_name = base.rsplit('.',1)[0] + '.xlsx'
                        excel_files.append((new_name, converted))
                    except Exception as conv_e:
                        # If conversion fails, add the original and allow read_excel_bytes_safe to attempt (may require xlrd)
                        st.warning(f"Conversion of {name} failed: {conv_e}. Will pass original bytes to reader (may need xlrd).")
                        excel_files.append((base, bts))
                else:
                    excel_files.append((base, bts))
            except Exception as read_e:
                st.warning(f"Could not read {name} inside ZIP: {read_e}")
    return excel_files

# ----------------- Existing helpers (header find, numeric parse, sheet totals) -----------------
def find_header_row(raw_df):
    nrows = min(10, len(raw_df))
    for i in range(nrows):
        row = raw_df.iloc[i].astype(str).str.lower().tolist()
        if any(('description' in c) or ('no of txns' in c) or ('debit' in c) or ('credit' in c) or ('transaction' in c) or ('total fee' in c) or ('total transaction' in c) for c in row):
            return i
    return 0

def extract_date_from_raw(raw):
    for col in raw.columns:
        for v in raw[col].astype(str).head(10):
            if v is None:
                continue
            m = date_regex.search(str(v))
            if m:
                return m.group(0)
    return ''

def normalize_cols(cols):
    return [str(c).strip().lower() for c in cols]

def safe_numeric(series):
    return pd.to_numeric(series.astype(str).str.replace(',','').str.replace('₹','').str.strip(), errors='coerce').fillna(0.0)

def extract_sheet_level_totals(raw, header_idx):
    res = {'sheet_fee_debit':0.0,'sheet_fee_credit':0.0,'sheet_transaction_debit':0.0,'sheet_transaction_credit':0.0}
    header_row = raw.iloc[header_idx].astype(str).tolist() if header_idx < len(raw) else []
    header_norm = [str(h).strip().lower() for h in header_row]
    cols = header_norm
    fee_idxs = [i for i,c in enumerate(cols) if 'total fee' in c or 'total fee amount' in c]
    txn_idxs = [i for i,c in enumerate(cols) if 'total transaction' in c or 'total transaction amount' in c or 'total transactionamount' in c]
    if not fee_idxs:
        fee_idxs = [i for i,c in enumerate(cols) if 'total' in c and 'fee' in c]
    if not txn_idxs:
        txn_idxs = [i for i,c in enumerate(cols) if 'total' in c and ('transaction' in c or 'transaction amount' in c)]
    data_row_idx = header_idx + 1
    if data_row_idx >= len(raw):
        return res
    data_row = raw.iloc[data_row_idx].tolist()

    def first_two_numbers_from(idx):
        nums=[]
        for j in range(idx, min(len(data_row), idx+6)):
            v = data_row[j]
            try:
                if v is None:
                    continue
                s = str(v).strip().replace(',','').replace('₹','')
                if re.match(r'^[A-Za-z\W_]+$', s):
                    continue
                num = float(s)
                nums.append(num)
                if len(nums)>=2:
                    break
            except Exception:
                m = re.findall(r'[-+]?\d*\.\d+|\d+', str(v).replace(',',''))
                if m:
                    try:
                        nums.append(float(m[0]))
                    except:
                        pass
                if len(nums)>=2:
                    break
        return nums

    for idx in fee_idxs:
        nums = first_two_numbers_from(idx)
        if nums:
            res['sheet_fee_debit'] += (nums[0] if len(nums)>=1 else 0.0)
            if len(nums)>=2:
                res['sheet_fee_credit'] += nums[1]
            break
    for idx in txn_idxs:
        nums = first_two_numbers_from(idx)
        if nums:
            res['sheet_transaction_debit'] += (nums[0] if len(nums)>=1 else 0.0)
            if len(nums)>=2:
                res['sheet_transaction_credit'] += nums[1]
            break

    # fallback: look for a date-start row and heuristically pick numbers if labels not found
    for r in range(0, min(6, len(raw))):
        first = str(raw.iloc[r,0]) if raw.shape[1]>0 else ''
        if first and date_regex.search(first):
            rowvals = raw.iloc[r].tolist()
            if not fee_idxs:
                nums=[]
                for j in range(1, min(len(rowvals),7)):
                    v = rowvals[j]
                    try:
                        if v is None:
                            continue
                        s = str(v).strip().replace(',','').replace('₹','')
                        if re.match(r'^[A-Za-z\W_]+$', s):
                            continue
                        nums.append(float(s))
                        if len(nums)>=2:
                            break
                    except:
                        m = re.findall(r'[-+]?\d*\.\d+|\d+', str(v).replace(',',''))
                        if m:
                            nums.append(float(m[0]))
                        if len(nums)>=2:
                            break
                if nums:
                    res['sheet_fee_debit'] = max(res['sheet_fee_debit'], nums[0])
                    if len(nums)>1:
                        res['sheet_fee_credit'] = max(res['sheet_fee_credit'], nums[1])
            if not txn_idxs:
                nums=[]
                for j in range(3, min(len(rowvals),9)):
                    v=rowvals[j] if j < len(rowvals) else None
                    try:
                        if v is None:
                            continue
                        s = str(v).strip().replace(',','').replace('₹','')
                        if re.match(r'^[A-Za-z\W_]+$', s):
                            continue
                        nums.append(float(s))
                        if len(nums)>=2:
                            break
                    except:
                        m = re.findall(r'[-+]?\d*\.\d+|\d+', str(v).replace(',',''))
                        if m:
                            nums.append(float(m[0]))
                        if len(nums)>=2:
                            break
                if nums:
                    res['sheet_transaction_debit'] = max(res['sheet_transaction_debit'], nums[0])
                    if len(nums)>1:
                        res['sheet_transaction_credit'] = max(res['sheet_transaction_credit'], nums[1])
            break
    return res

# ----------------- Processing (same as earlier, using conversion) -----------------
def process_excel_bytes(file_bytes, filename, fee_keywords, txn_keywords, debug=False):
    totals = {
        'file': filename, 'date': '',
        'total_fee_debit': 0.0, 'total_fee_credit': 0.0,
        'total_transaction_debit': 0.0, 'total_transaction_credit': 0.0,
        'sheet_fee_debit': 0.0, 'sheet_fee_credit': 0.0,
        'sheet_transaction_debit': 0.0, 'sheet_transaction_credit': 0.0
    }
    try:
        sheets = read_excel_bytes_safe(file_bytes, filename)
    except RuntimeError as re_err:
        raise re_err
    except Exception as e:
        st.warning(f"Could not parse {filename}: {e}")
        return totals

    for sheet_name, raw in sheets.items():
        if raw is None or raw.shape[0] == 0:
            continue
        header_idx = find_header_row(raw)
        if header_idx >= len(raw) - 1:
            header_idx = 0
        header_vals = raw.iloc[header_idx].fillna('').astype(str).tolist()
        df = raw.iloc[header_idx+1:].copy().reset_index(drop=True)
        df.columns = header_vals

        # extract date
        if not totals['date']:
            d = extract_date_from_raw(raw)
            if d:
                totals['date'] = d

        # extract sheet-level totals and add to totals
        sheet_totals = extract_sheet_level_totals(raw, header_idx)
        for k,v in sheet_totals.items():
            totals[k] = totals.get(k,0.0) + v

        cols = normalize_cols(df.columns)
        df.columns = cols

        # header heuristics to avoid No of Txns selection
        desc_col = next((c for c in cols if 'description' in c), None)
        ignore_tokens = ['no of txn', 'no of txns', 'no.of.txn', 'no.of.txns', 'no of transactions', 'count', 'no. of txns']
        def header_is_count(h):
            if not h:
                return False
            hl = h.lower()
            return any(tok in hl for tok in ignore_tokens)

        debit_col = next((c for c in cols if ('debit' in c or 'transaction amount' in c or 'amount' in c) and not header_is_count(c)), None)
        credit_col = next((c for c in cols if ('credit' in c or 'transaction amount' in c or 'amount' in c) and not header_is_count(c) and c != debit_col), None)

        if debit_col is None or credit_col is None:
            numeric_candidates = []
            for c in cols:
                if header_is_count(c):
                    continue
                try:
                    series = pd.to_numeric(df[c], errors='coerce').dropna()
                    if len(series) > 0:
                        numeric_candidates.append((c, series.abs().mean(), len(series)))
                except Exception:
                    continue
            numeric_candidates.sort(key=lambda x: (x[1], x[2]), reverse=True)
            if numeric_candidates:
                if debit_col is None:
                    debit_col = numeric_candidates[0][0]
                if credit_col is None:
                    for cand in numeric_candidates:
                        if cand[0] != debit_col:
                            credit_col = cand[0]
                            break

        if desc_col is None and len(cols) > 0:
            desc_col = cols[0]
        if debit_col is None:
            debit_col = cols[1] if len(cols) > 1 else cols[0]
        if credit_col is None:
            credit_col = cols[2] if len(cols) > 2 else None
            if credit_col is None:
                df['__credit__'] = 0.0
                credit_col = '__credit__'

        try:
            subset = df[[desc_col, debit_col, credit_col]].copy()
        except Exception:
            continue
        subset.columns = ['description', 'debit', 'credit']
        subset['debit'] = safe_numeric(subset['debit'])
        subset['credit'] = safe_numeric(subset['credit'])

        desc_lower = subset['description'].astype(str).str.lower()

        fee_pattern = '|'.join([re.escape(k.strip().lower()) for k in fee_keywords if k.strip()]) or 'fee'
        txn_phrases = [k.strip().lower() for k in txn_keywords if k.strip()]
        txn_pattern = '|'.join([r'\b' + re.escape(p) + r'\b' for p in txn_phrases]) if txn_phrases else r'\btransaction amount\b'

        fee_mask = desc_lower.str.contains(fee_pattern, na=False)
        count_desc_mask = desc_lower.str.contains('no of txn|no of txns|no.of.txn|count|no. of txns', regex=True, na=False)
        txn_mask = desc_lower.str.contains(txn_pattern, regex=True, na=False) & (~fee_mask) & (~count_desc_mask)

        totals['total_fee_debit'] += subset.loc[fee_mask, 'debit'].sum()
        totals['total_fee_credit'] += subset.loc[fee_mask, 'credit'].sum()
        totals['total_transaction_debit'] += subset.loc[txn_mask, 'debit'].sum()
        totals['total_transaction_credit'] += subset.loc[txn_mask, 'credit'].sum()

        if debug:
            debug_df = subset.copy()
            debug_df['is_fee'] = fee_mask
            debug_df['is_transaction_amount'] = txn_mask
            st.write(f"Debug preview for file: {filename}, sheet: {sheet_name}")
            st.dataframe(debug_df.head(200))

    if not totals['date']:
        m = date_regex.search(filename)
        if m:
            totals['date'] = m.group(0)
    return totals

# ----------------- UI -----------------
st.sidebar.header("Options")
fee_kw_text = st.sidebar.text_input("Fee keywords (comma separated)", value="fee, switching fee, psp fee")
txn_kw_text = st.sidebar.text_input("Transaction keywords (comma separated)", value="transaction amount, approved transaction amount")
fee_kw = [x.strip() for x in fee_kw_text.split(',') if x.strip()]
txn_kw = [x.strip() for x in txn_kw_text.split(',') if x.strip()]

debug_inspect = st.sidebar.checkbox("Show row-level classification (debug)", value=False)

st.info("Upload a ZIP file containing the Excel sheets or upload multiple .xlsx/.xls files directly.")

uploaded = st.file_uploader("Upload ZIP or Excel files", type=["zip","xlsx","xls"], accept_multiple_files=True)

if uploaded:
    file_bytes_list = []
    for u in uploaded:
        name = u.name
        try:
            data = u.read()
        except Exception as e:
            st.warning(f"Could not read uploaded file {name}: {e}")
            continue

        if name.lower().endswith('.zip'):
            extracted = extract_excel_files_from_zip(data)
            if not extracted:
                st.warning(f"No Excel files found inside {name}.")
            else:
                file_bytes_list.extend(extracted)
        else:
            base = name.split('/')[-1]
            if base.startswith('._') or base.startswith('~$'):
                st.warning(f"Skipping temporary file {name}.")
                continue
            # if direct .xls uploaded, convert to .xlsx bytes before appending
            if base.lower().endswith('.xls'):
                try:
                    conv = convert_xls_to_xlsx_bytes(data)
                    new_name = base.rsplit('.',1)[0] + '.xlsx'
                    file_bytes_list.append((new_name, conv))
                except Exception as conv_e:
                    st.warning(f"Conversion of {base} failed: {conv_e}. Will attempt to process original file.")
                    file_bytes_list.append((base, data))
            else:
                file_bytes_list.append((name, data))

    if not file_bytes_list:
        st.warning("No readable Excel files found in the upload.")
    else:
        if st.button("Process files"):
            progress = st.progress(0)
            results = []
            total_files = len(file_bytes_list)
            for i, (fname, fbytes) in enumerate(file_bytes_list):
                try:
                    totals = process_excel_bytes(fbytes, fname, fee_kw, txn_kw, debug=debug_inspect)
                    results.append(totals)
                except RuntimeError as e:
                    st.error(f"Error processing {fname}: {e}")
                except Exception as e:
                    st.warning(f"Could not read {fname}: {e}")
                progress.progress(int((i + 1) / total_files * 100))

            if results:
                out_df = pd.DataFrame(results)

                # fill date column with modal/common date if empty
                out_df['date'] = out_df['date'].fillna('').astype(str).str.strip()
                non_empty_dates = out_df.loc[out_df['date'] != '', 'date']
                common_date = ''
                if not non_empty_dates.empty:
                    try:
                        common_date = non_empty_dates.mode().iloc[0]
                    except:
                        common_date = non_empty_dates.iloc[0]
                else:
                    for fname in out_df['file'].astype(str):
                        m = date_regex.search(fname)
                        if m:
                            common_date = m.group(0)
                            break
                if common_date:
                    out_df['date'] = out_df['date'].replace('', common_date)

                # round numeric columns to 2 decimals
                num_cols = [
                    'total_fee_debit','total_fee_credit',
                    'total_transaction_debit','total_transaction_credit',
                    'sheet_fee_debit','sheet_fee_credit',
                    'sheet_transaction_debit','sheet_transaction_credit'
                ]
                for c in num_cols:
                    if c in out_df.columns:
                        out_df[c] = pd.to_numeric(out_df[c], errors='coerce').fillna(0.0).round(2)
                    else:
                        out_df[c] = 0.0

                # show per-file results
                st.subheader("Per-file summary (computed and sheet-level)")
                st.dataframe(out_df.style.format({c:'{:.2f}' for c in num_cols}), use_container_width=True)

                # overall totals
                overall = {
                    'file': 'ALL_FILES_SUM',
                    'date': common_date if common_date else '',
                    'total_fee_debit': round(out_df['total_fee_debit'].sum(), 2),
                    'total_fee_credit': round(out_df['total_fee_credit'].sum(), 2),
                    'total_transaction_debit': round(out_df['total_transaction_debit'].sum(), 2),
                    'total_transaction_credit': round(out_df['total_transaction_credit'].sum(), 2),
                    'sheet_fee_debit': round(out_df['sheet_fee_debit'].sum(), 2),
                    'sheet_fee_credit': round(out_df['sheet_fee_credit'].sum(), 2),
                    'sheet_transaction_debit': round(out_df['sheet_transaction_debit'].sum(), 2),
                    'sheet_transaction_credit': round(out_df['sheet_transaction_credit'].sum(), 2)
                }
                st.subheader("Overall totals (all uploaded files)")
                st.table(pd.DataFrame([overall]).T.rename(columns={0:'value'}))

                csv_bytes = out_df.to_csv(index=False, float_format='%.2f').encode('utf-8')
                st.download_button("Download summary CSV", data=csv_bytes, file_name="settlement_summary.csv", mime="text/csv")
                st.success("Done.")
            else:
                st.warning("No results produced (all files may have failed reading).")

#st.markdown("---")
#st.markdown("Notes: .xls files are converted to .xlsx automatically (requires pyexcel + pyexcel-xls). If conversion fails, the app will attempt to read the original .xls (which requires xlrd). Use debug checkbox to inspect matched rows.")
